from openpyxl import load_workbook

workbook = load_workbook('./sendtank_customer.xlsx', user_iterators=True)
first_sheet = workbook.get_sheet_names()[0]
worksheet = workbook.get_sheet_by_name(first_sheet)

for row in worksheet.iter_rows():
    print(row)

for cell in row:
    print(cell)
